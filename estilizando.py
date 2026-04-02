# # BIBLIOTECAS IMPORTADAS
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill, Alignment, Border, Side
from datetime import datetime

# DEFININDO ATALHOS
agora = datetime.now()
arquivo = Workbook()
planilha_ativa = arquivo.active
planilha_ativa.title = 'PRODUTOS'

# # ESCREVENDO
planilha_ativa['A1'] = 'RELATÓRIO DE VENDAS'
planilha_ativa.append(['PRODUTO', 'PREÇO', 'ESTOQUE', 'DATA'])
planilha_ativa.append(['Camiseta', 29.99, 20, agora])
planilha_ativa.append(['Bermuda', 47.50, 25, agora])
planilha_ativa.append(['Tênis', 200, 12, agora])

# # MESCLANDO CÉLULAS COM: .merge_cells(intervalo)
planilha_ativa.merge_cells('A1:D1')

# # DEFININDO O ESTILO DA CÉLULA MESCLADA (TÍTULO GERAL) COM MANUALMENTE: Font(type/color) | PatternFill(fgColor/fill_type) | Alignment(direction)
planilha_ativa['A1'].font = Font(bold=True, color='FFFFFF')
planilha_ativa['A1'].fill = PatternFill(fgColor='1D12B6', fill_type='solid')
planilha_ativa['A1'].alignment = Alignment(horizontal='center')


# # DEFININDO O ESTILO DO CABEÇALHO (TÍTULO DA COLUNA) COM VARIÁVEIS DE ATALHOS E LOOP FOR + POSIÇÃO: Font | PatternFill | Alignment
conteudo_celula = Font(bold=True, color='FFFFFF')
fundo_celula = PatternFill(fgColor='464555', fill_type='solid')
alinhamento = Alignment(horizontal='center')

for celula in planilha_ativa[2]:
    celula.font = conteudo_celula
    celula.fill = fundo_celula
    celula.alignment = alinhamento

# # DEFININDO E APLICANDO ESTILO DE BORDA COM VARIÁVEIS DE ATALHO, LOOP FOR E iter_rows(): Border(local) e Side(estilo)
estilo_linha = Side('thick')
lado_linha = Border(left=estilo_linha, right=estilo_linha, top=estilo_linha, bottom=estilo_linha)

for linha in planilha_ativa.iter_rows():
    for celula in linha:
        celula.border = lado_linha

# # FORMANTO VALORES EM REAIS(R$) COM LOOP FOR E CONCICIONAL .row <=: number_format = "R$ #,##0.00"
for celula in planilha_ativa['B']:
    if celula.row <= 2:
        continue
    celula.number_format = "R$ #,##0.00" "0.0%"
    celula.alignment = Alignment(horizontal='center')

for celula in planilha_ativa['C']:
    if celula.row <= 2:
        continue
    celula.alignment = Alignment(horizontal='center')

# # FORMATANDO DATAS COM LOOP FOR E CONCICIONAL .row <=: number_format = "DD/MM/YYYY"
for celula in planilha_ativa['D']:
    if celula.row <= 2:
        continue
    celula.number_format = "DD/MM/YYYY"

# # CRIANDO O ARQUIVO .xlsx COM AS ALTERAÇÕES REALIZADAS
arquivo.save('relatorio_vendas.xlsx')