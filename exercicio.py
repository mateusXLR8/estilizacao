from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


arquivo = Workbook()
planilha = arquivo.active
planilha.title = 'Leituras'


planilha['A1'] = 'Diário de Leituras – Março 2026'
planilha.merge_cells('A1:D1')


planilha.append(['Livro', 'Autor', 'Data de Início', 'Progresso (%)'])


planilha.append(['Dom Casmurro', 'Machado de Assis', datetime(2026, 3, 1), 0.5])
planilha.append(['1984', 'George Orwell', datetime(2026, 3, 5), 0.3])
planilha.append(['O Hobbit', 'J.R.R. Tolkien', datetime(2026, 3, 10), 0.8])
planilha.append(['A Revolução dos Bichos', 'George Orwell', datetime(2026, 3, 15), 0.6])


planilha['A1'].font = Font(bold=True, color='FFFFFF', size=14)
planilha['A1'].fill = PatternFill(fgColor='1F497D', fill_type='solid')
planilha['A1'].alignment = Alignment(horizontal='center')


fonte_cab = Font(bold=True, color='FFFFFF')
fundo_cab = PatternFill(fgColor='4f4f4f', fill_type='solid')
alinhamento_centro = Alignment(horizontal='center')

for celula in planilha[2]:
    celula.font = fonte_cab
    celula.fill = fundo_cab
    celula.alignment = alinhamento_centro


borda_fina = Side(style='thin')
borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

for linha in planilha.iter_rows():
    for celula in linha:
        celula.border = borda


for i, linha in enumerate(planilha.iter_rows(min_row=3), start=3):

    
    cor = 'FFFFFF' if i % 2 == 1 else 'D9D9D9'
    for celula in linha:
        celula.fill = PatternFill(fgColor=cor, fill_type='solid')

    
    linha[0].alignment = Alignment(horizontal='left')   # Livro
    linha[1].alignment = Alignment(horizontal='left')   # Autor
    linha[2].alignment = Alignment(horizontal='center') # Data
    linha[3].alignment = Alignment(horizontal='right')  # Progresso

    
    linha[2].number_format = 'DD/MM/YYYY'

    
    linha[3].number_format = '0.0%'


arquivo.save('diario_leituras.xlsx')
